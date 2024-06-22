from django.urls import reverse
import calendar
from django.db.models.functions import ExtractYear, ExtractMonth
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment
import openpyxl
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.utils.dateparse import parse_time
# for rendering templates and redirecting
from django.shortcuts import render, redirect, HttpResponse
# for authentication and login
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator  # for login required decorator
from django.utils import timezone
from .managers import create_current_day_attendances
from datetime import datetime
from .models import *
from django.db.models import Sum
from django.contrib import messages
from django.shortcuts import render, get_object_or_404
from .forms import *
from .managers import *
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from django.http import JsonResponse
from django.db.models import F, Count, Case, When, Value, IntegerField
from django.db.models.functions import ExtractMonth, ExtractYear
from PIL import Image
from django.core.files.base import ContentFile
from io import BytesIO
from django.core.files.uploadedfile import SimpleUploadedFile
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.template.loader import render_to_string
from django.utils.crypto import get_random_string
from django.core.mail import send_mail
from .models import PasswordResetRequest
from django.shortcuts import redirect
import os
from geopy.geocoders import Nominatim
import datetime
from django.db.models import Sum


geolocator = Nominatim(user_agent="urbanplace_hr")

@login_required
def home(request):
    # Get the current date
    today = timezone.now()
    current_date = today.date()

    # Calculate the start and end date of the current month
    current_month = today.month
    current_year = today.year
    start_date = timezone.datetime(current_year, current_month, 1).date()
    if current_month == 12:
        end_date = timezone.datetime(current_year + 1, 1, 1).date()
    else:
        end_date = timezone.datetime(current_year, current_month + 1, 1).date()

    # Get all employees in the same company as the current user
    employees = Employee.objects.filter(company=request.user.employee.company)

    # Get all present attendance records for the current day
    present_attendance_records = Attendance.objects.filter(
        date=current_date, is_present=True)

    # Calculate the number of present employees
    total_present = present_attendance_records.count()

    # Calculate the number of absent employees
    total_employees = employees.count()
    total_absent = total_employees - total_present

    # Get counts for present, absent, late, and early for the current day
    daily_counts = Attendance.objects.filter(date=current_date).aggregate(
        total_present=Count(
            Case(When(is_present=True, then=Value(1)), output_field=IntegerField())),
        total_absent=Count(
            Case(When(is_absent=True, then=Value(1)), output_field=IntegerField())),
        total_late=Count(
            Case(When(is_late=True, then=Value(1)), output_field=IntegerField())),
        total_early=Count(
            Case(When(is_early=True, then=Value(1)), output_field=IntegerField())),
    )

    # Query to get monthly data
    employee = Attendance.objects.filter(
        date__gte=start_date,
        date__lt=end_date,
        employee__user=request.user
    ).values(
        "employee__user__email",  # Grouping by employee email
        "employee__id",            # Also retrieving employee ID
    ).annotate(
        year=ExtractYear('date'),   # Extract year from date
        month=ExtractMonth('date'),  # Extract month from date
        total_records=Count('id'),   # Count of total records for each employee
        present=Count(Case(When(is_present=True, then=Value(1)),
                      output_field=IntegerField())),  # Count of present days
        absent=Count(Case(When(is_absent=True, then=Value(1)),
                     output_field=IntegerField())),    # Count of absent days
        half_day=Count(Case(When(is_half_day=True, then=Value(
            1)), output_field=IntegerField())),  # Count of half-day attendance
        leave=Count(Case(When(is_leave=True, then=Value(1)),
                    output_field=IntegerField())),      # Count of leave days
        # Count of late attendance
        late=Count(Case(When(is_late=True, then=Value(1)),
                   output_field=IntegerField())),
        # Count of early departures
        early=Count(Case(When(is_early=True, then=Value(1)),
                    output_field=IntegerField())),
    ).order_by('employee__user__email')

    context = {
        'employee': employee.first(),
        'daily_counts': daily_counts,
        'present_attendance_records': present_attendance_records,
        'total_absent': total_absent,
        'total_employees': total_employees,
        'total_present': total_present,
    }

    return render(request, 'home.html', context)


# Login view
def login_view(request):
    print(request.method)   # print request method
    context = {}  # prepare context for error message

    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')

        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            user = None

        # Check if user exists and authenticate with provided password
        if user is not None:
            authenticated_user = authenticate(
                request, username=user.username, password=password)

            if authenticated_user is not None:
                login(request, authenticated_user)   # log in the user
                return redirect('/', permanent=True)  # redirect to home page
            else:
                # incorrect password error
                context["error_message"] = 'Invalid email or password.'
        else:
            # user not found error
            context["error_message"] = 'Invalid email or password.'

    # Redirect if the user is already authenticated
    if request.user.is_authenticated:
        return redirect('/', permanent=True)

    # render login page with error message if any
    return render(request, 'login.html', context)

# Logout View


def logout_view(request):  # for logout
    logout(request)
    return redirect('login')  # redirect to login page


# compressed images before saving View

def compress_image(image, quality=50):
    """
    Compresses an image stored in a Django ImageField and saves the compressed version.

    Args:
        image: Django ImageField instance containing the image.
        quality (int): Quality of the compressed image (0-100). Lower values mean higher compression.
    """
    if not image:
        return

    img = Image.open(image)

    # Create a BytesIO buffer to temporarily store the compressed image
    img_buffer = BytesIO()

    # Compress and save the image to the buffer
    img.save(img_buffer, format='webp', quality=quality)

    # Set the buffer's file pointer to the beginning
    img_buffer.seek(0)
    return img_buffer

# Check In View


@login_required
def check_in(request):
    employee = Employee.objects.get(user=request.user)
    if request.method == 'POST':
        attn_date = datetime.date.today()
        try:
            attn = Attendance.objects.get(employee=employee, date=attn_date)
        except Attendance.DoesNotExist:
            # If attendance entry for today doesn't exist, create a new one
            attn = Attendance(employee=employee, date=attn_date)
            attn.save()

        image = request.FILES.get("clock_in_image")
        if image:
            # Compress the image
            compressed_image = compress_image(image, 50)

            # Get bytes from the BytesIO object
            compressed_image.seek(0)
            image_bytes = compressed_image.read()

            # Create an InMemoryUploadedFile object from the compressed image data
            uploaded_image = SimpleUploadedFile(os.path.basename(
                image.name), image_bytes, content_type='image/webp')

            # Save clock in details
            attn.clock_in = datetime.datetime.now().time()
            attn.clock_in_image = uploaded_image
            location_data = attn.coords
            location_data["check_in_loaction"] = {
                "lat": request.POST.get("lat"), "lng": request.POST.get("lng")}
            location_data["check_in_address"]=geolocator.reverse(f"{request.POST.get('lat')},{request.POST.get('lng')}").address
            attn.coords = location_data
            attn.save()

    return redirect('/')

# Check Out View


@login_required
def check_out(request):
    employee = Employee.objects.get(user=request.user)
    if request.method == 'POST':
        attn_date = datetime.date.today()
        try:
            attn = Attendance.objects.get(employee=employee, date=attn_date)
        except Attendance.DoesNotExist:
            # If attendance entry for today doesn't exist, create a new one
            attn = Attendance(employee=employee, date=attn_date)
            attn.save()

        image = request.FILES.get("clock_out_image")
        if image:
            # Compress the image
            compressed_image = compress_image(image, 50)

            # Get bytes from the BytesIO object
            compressed_image.seek(0)
            image_bytes = compressed_image.read()

            # Create an InMemoryUploadedFile object from the compressed image data
            uploaded_image = SimpleUploadedFile(os.path.basename(
                image.name), image_bytes, content_type='image/webp')

        # Update clock-out details
        attn.clock_out = datetime.datetime.now().time()
        attn.clock_out_image = uploaded_image
        location_data = attn.coords
        location_data["check_out_loaction"] = {
            "lat": request.POST.get("lat"), "lng": request.POST.get("lng")}
        location_data["check_out_address"]=geolocator.reverse(f"{request.POST.get('lat')},{request.POST.get('lng')}").address
        attn.coords = location_data
        attn.save()

    return redirect('/')

# Profile View


@login_required
def profile_setting(request):
    try:

        user_employee = request.user.employee

        if request.method == 'POST':

            user_employee.user.first_name = request.POST.get(
                'full_name').split(" ")[0]
            user_employee.user.last_name = request.POST.get(
                'full_name').split(" ")[-1]
            user_employee.user.email = request.POST.get('email')
            user_employee.phone_number = request.POST.get('phone_number')
            user_employee.city = request.POST.get('city')
            user_employee.state = request.POST.get('state')
            user_employee.zip_code = request.POST.get('zip_code')
            user_employee.country = request.POST.get('country')
            user_employee.department = request.POST.get('department')

            if request.FILES.get('profile_picture'):
                user_employee.image = request.FILES['profile_picture']

            user_employee.user.save()
            user_employee.save()

            return redirect('/')

    except Employee.DoesNotExist:

        pass

    return render(request, "profile_setting.html", {'user': user_employee})

# Company Profile view


@login_required
def company_profile(request):
    company = Company.objects.first()  # Assuming there's only one company profile

    if request.method == 'POST':
        name = request.POST.get('name')
        est_year = request.POST.get('est_year')
        about = request.POST.get('about')
        location = request.POST.get('location')

        # Get logo file from request.FILES
        logo_file = request.FILES.get('logo')

        # Update company instance with form data
        company.name = name
        company.est_year = est_year
        company.about = about
        company.location = location

        # If logo file is provided, save it
        if logo_file:
            company.logo = logo_file

        # Save updated company instance
        company.save()

        # Redirect to the same page after saving
        return redirect('company_profile')

    return render(request, 'company_profile.html', {'company': company})

# Edit Company Profile view


@login_required
def edit_company(request):
    if not request.user.is_superuser:
        messages.error(
            request, "You don't have permission to access this page.")
        return redirect('company_profile')

    company = Company.objects.first()  # Assuming there's only one company profile

    if request.method == 'POST':
        form = CompanyForm(request.POST, request.FILES, instance=company)
        if form.is_valid():
            form.save()
            messages.success(request, 'Company details updated successfully.')
            return redirect('company_profile')
    else:
        form = CompanyForm(instance=company)

    return render(request, 'edit_company.html', {'form': form})

# Settings View


@login_required
def settings_view(request):
    holidays = Holiday.objects.all()
    return render(request, "settings.html", {'holidays': holidays})

# Employee View


@login_required
def employees_view(request):
    employees = Employee.objects.filter(company=request.user.employee.company)
    print(employees)

    shifts = []
    for i in Shift.objects.all():
        print(type(i.start_time))
        data = {
            "id": i.id,
            "name": i.title,
            "start_time": i.start_time,
            "end_time": i.end_time,
            "full_day_hours": i.full_day_hours,
            "half_day_hours": i.half_day_hours,
            "penalty":i.CompanyPenality
        }
        data["weekoffs"] = []
        if not i.monday:
            data["weekoffs"].append("Monday")
        if not i.tuesday:
            data["weekoffs"].append("Tuesday")
        if not i.wednesday:
            data["weekoffs"].append("Wednesday")
        if not i.thursday:
            data["weekoffs"].append("Thursday")
        if not i.friday:
            data["weekoffs"].append("Friday")
        if not i.saturday:
            data["weekoffs"].append("Saturday")
        if not i.sunday:
            data["weekoffs"].append("Sunday")
        shifts.append(data)

    context = {
        "employees": employees,
        "shifts": shifts
    }

    return render(request, 'employees.html', context)


# employee details and assign shift views
@login_required
def employee_details_view(request, employee_id=None):
    search_text = request.GET.get('search', '')
    
    employees = Employee.objects.filter(company=request.user.employee.company)

    if search_text:
        employees = employees.filter(
            Q(user__first_name__icontains=search_text) | 
            Q(user__last_name__icontains=search_text)
        )

    if employee_id is None:
        employee = employees.first()
    else:
        employee = get_object_or_404(employees, user__id=employee_id)

    # Retrieve all shifts
    shifts = []
    for i in Shift.objects.all():
        data = {
            "id": i.id,
            "name": i.title,
            "start_time": i.start_time,
            "end_time": i.end_time,
            "full_day_hours": i.full_day_hours,
            "half_day_hours": i.half_day_hours,
            "penalty": i.CompanyPenality
        }
        data["weekoffs"] = []
        if not i.monday:
            data["weekoffs"].append("Monday")
        if not i.tuesday:
            data["weekoffs"].append("Tuesday")
        if not i.wednesday:
            data["weekoffs"].append("Wednesday")
        if not i.thursday:
            data["weekoffs"].append("Thursday")
        if not i.friday:
            data["weekoffs"].append("Friday")
        if not i.saturday:
            data["weekoffs"].append("Saturday")
        if not i.sunday:
            data["weekoffs"].append("Sunday")
        shifts.append(data)

    # Handle form submission
    if request.method == 'POST':
        # Extract form data from POST request
        phone_number = request.POST.get('phone_number')
        date_of_birth = request.POST.get('date_of_birth')
        city = request.POST.get('city')
        state = request.POST.get('state')
        zip_code = request.POST.get('zip_code')
        country = request.POST.get('country')
        department = request.POST.get('department')
        position = request.POST.get('position')
        report_to = request.POST.get('report_to')
        print(report_to)

        # Assign report_to field if provided
        if report_to == 'None':
            report_to = None  # Convert string 'None' to Python None

        if report_to:
            report_to_employee = get_object_or_404(Employee, id=report_to)
            employee.report_to = report_to_employee
        else:
            employee.report_to = None

        # Assign form data to employee fields
        employee.date_of_birth = date_of_birth if date_of_birth else ""
        employee.city = city if city else ""
        employee.state = state if state else ""
        employee.country = country if country else ""
        employee.department = department if department else ""
        employee.position = position if position else ""
        employee.save()

    # Retrieve attendance records based on the employee ID if provided
    attendances = Attendance.objects.filter(employee=employee)

    # Leave-related logic
    current_month = datetime.datetime.now().month
    current_year = datetime.datetime.now().year
    leaves = LeaveRequest.objects.filter(
        employee__id=employee.id,
        from_date__month=current_month,
        from_date__year=current_year,
    )

    total_leaves_taken = leaves.aggregate(
        total=Sum('number_of_leaves'))['total'] or 0
    company_policy = CompanyPolicy.objects.filter(
        company=employee.company).first()
    total_leaves_allowed = company_policy.total_leaves if company_policy else 0
    sick_leave_balance = company_policy.sick_leaves if company_policy else 0
    paid_leave_balance = company_policy.paid_leaves if company_policy else 0
    pending_leaves = total_leaves_allowed - total_leaves_taken

    context = {
        "employee": employee,
        "employees": employees,
        "shifts": shifts,
        "attendances": attendances,
        "total_leaves_allowed": total_leaves_allowed,
        "sick_leave_balance": sick_leave_balance,
        "paid_leave_balance": paid_leave_balance,
        "leaves": leaves,
        "total_leaves_taken": total_leaves_taken,
        "pending_leaves": pending_leaves,
        "search_text": search_text
    }
    return render(request, 'employee_details.html', context)





@login_required
def shift_assign_schedule(request):
    try:
        if request.method == 'POST':
            shiftassign = request.POST.get("shiftassign")
            emp = request.user.employee
            shift = Shift.objects.get(id=shiftassign)
            emp.shift = shift
            emp.save()
            return redirect(employee_details_view)
    except Exception as e:
        print("An error occurred:", e)
    return None


@login_required
def attendance_view(request):

    if request.method == 'POST':
        post = request.POST
        print(post)
        shift = Shift(
            company=request.user.employee.company, name=post.get("name"), start_time=post.get("start_time"), end_time=post.get("end_time"), full_day_hours=post.get("full_day_hours"), half_day_hours=post.get("half_day_hours"))
        for i in post.get("weekdays"):
            if i == "Monday":
                shift.monday = True
            if i == "Tuesday":
                shift.tuesday = True
            if i == "Wednesday":
                shift.wednesday = True
            if i == "Thursday":
                shift.thursday = True
            if i == "Friday":
                shift.friday = True
            if i == "Saturday":
                shift.saturday = True
            if i == "Sunday":
                shift.sunday = True
        shift.save()
    shifts = []
    for i in Shift.objects.all():
        data = {
            "id": i.id,
            "name": i.title,
            "start_time": i.start_time,
            "end_time": i.end_time,
            "full_day_hours": i.full_day_hours,
            "half_day_hours": i.half_day_hours,
            "penalty":i.CompanyPenality
        }
        data["weekoffs"] = []
        if not i.monday:
            data["weekoffs"].append("Monday")
        if not i.tuesday:
            data["weekoffs"].append("Tuesday")
        if not i.wednesday:
            data["weekoffs"].append("Wednesday")
        if not i.thursday:
            data["weekoffs"].append("Thursday")
        if not i.friday:
            data["weekoffs"].append("Friday")
        if not i.saturday:
            data["weekoffs"].append("Saturday")
        if not i.sunday:
            data["weekoffs"].append("Sunday")
        shifts.append(data)
    return render(request, 'attendance.html', {"shifts": shifts})


@login_required
def edit_shift(request, shift_id):
    shift = get_object_or_404(Shift, pk=shift_id)
    if request.method == 'POST':
        form = ShiftForm(request.POST, instance=shift)
        if form.is_valid():
            form.save()
            return redirect('attendance')
    else:
        form = ShiftForm(instance=shift)
    return render(request, 'edit_shift.html')


@login_required
def holiday_list(request):
    holidays = Holiday.objects.all()
    return render(request, 'settings.html', )


def delete_shift(request, shift_id):
    shift = get_object_or_404(Shift, pk=shift_id)
    shift.delete()
    return redirect('attendance')


def delete_holiday(request, holiday_id):
    holiday = get_object_or_404(Holiday, pk=holiday_id)
    holiday.delete()
    return redirect('settings')


def calendar_view(request, employee_id=None):
    # Filter attendance records based on the employee ID if provided
    if employee_id is not None:
        attendances = Attendance.objects.filter(employee__id=employee_id)
    else:
        # If employee ID is not provided, retrieve all attendance records
        attendances = Attendance.objects.all()

    return render(request, 'calendar.html', {'attendances': attendances})


@login_required
def submit_attendance_request(request):
    if request.method == 'POST':
        # Default to 'Check-in' if not provided
        request_type = request.POST.get('request_type', default='Check-in')
        description = request.POST.get('description')
        date = request.POST.get('date')
        time_str = request.POST.get('time')

        # Parse and format the time string
        if time_str:
            time_obj = parse_time(time_str)
            if time_obj:
                formatted_time = time_obj.strftime('%H:%M:%S')
            else:
                messages.error(request, 'Invalid time format.')
                return redirect('submit_attendance_request')
        else:
            formatted_time = None  # Handle case where time might not be provided

        # Create an AttendanceRequest object and save it to the database
        attendance_request = AttendanceRequest.objects.create(
            request_type=request_type,
            description=description,
            employee=request.user.employee,
            date=date,
            time=formatted_time
        )

        messages.success(
            request, 'Your attendance request has been submitted successfully.')
        return redirect('/')

    # Default to 'check_in' for GET requests if request_type is not passed in query parameters
    request_type = request.GET.get('request_type', 'check_in')
    return render(request, 'submit_attendance_request.html', {'request_type': request_type})


@login_required
def submit_leave_request(request):
    if request.method == 'POST':
        request_type = request.POST.get('request_type')
        description = request.POST.get('description')
        from_date = request.POST.get('from_date')
        to_date = request.POST.get('to_date')

        # Create a LeaveRequest object and save it to the database
        leave_request = LeaveRequest.objects.create(
            request_type=request_type,
            description=description,
            from_date=from_date,
            to_date=to_date,
            employee=request.user.employee
        )
        messages.success(
            request, 'Your leave request has been submitted successfully.')
        return redirect('/')

    return render(request, 'submit_leave_request.html')


def add_holiday(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        date = request.POST.get('date')
        holiday_type = request.POST.get('holiday_type')

        if name and date and holiday_type:
            holiday = Holiday(
                name=name,
                date=date,
                holiday_type=holiday_type,
            )
            holiday.save()
            return redirect('settings')
    return render(request, 'add_holiday.html')


def add_shift(request):
    company_penalties = CompanyPenality.objects.all()  # Retrieve all company penalties

    if request.method == 'POST':
        # Retrieve data from the POST request
        shift_name = request.POST.get('title')
        start_time = request.POST.get('start_time')
        end_time = request.POST.get('end_time')
        full_day_hours = request.POST.get('full_day_hours')
        half_day_hours = request.POST.get('half_day_hours')
        start_time_buffer = request.POST.get('start_time_buffer')
        end_time_buffer = request.POST.get('end_time_buffer')
        working_days = request.POST.getlist('working_days')
        company_penality_id = request.POST.get('company_penality')

        # Validate if shift_name is provided
        if shift_name:
            # Create a new Shift object with the retrieved data
            shift = Shift(
                title=shift_name,
                start_time=start_time,
                end_time=end_time,
                full_day_hours=full_day_hours,
                half_day_hours=half_day_hours,
                start_time_buffer=start_time_buffer,
                end_time_buffer=end_time_buffer,
            )

            # Set the boolean fields (working days) based on the POST data
            shift.monday = 'monday' in working_days
            shift.tuesday = 'tuesday' in working_days
            shift.wednesday = 'wednesday' in working_days
            shift.thursday = 'thursday' in working_days
            shift.friday = 'friday' in working_days
            shift.saturday = 'saturday' in working_days
            shift.sunday = 'sunday' in working_days

            # Set the CompanyPenality if selected
            if company_penality_id:
                shift.CompanyPenality_id = company_penality_id

            # Save the shift object
            shift.save()

            # Redirect to appropriate URL after successful submission
            return redirect('attendance')  # Adjust 'attendance' to your desired URL name

    # If GET request or form not valid, render the form again with company penalties
    return render(request, 'add_shift.html', {'company_penalties': company_penalties})


def leaves_view(request, employee_id=None):
    current_month = datetime.datetime.now().month
    current_year = datetime.datetime.now().year
    leaves = LeaveRequest.objects.filter(
        employee__id=employee_id,
        from_date__month=current_month,
        from_date__year=current_year,
    )

    total_leaves_taken = leaves.aggregate(
        total=Sum('number_of_leaves'))['total'] or 0
    employee = get_object_or_404(Employee, id=employee_id)
    company_policy = CompanyPolicy.objects.filter(
        company=employee.company).first()
    total_leaves_allowed = company_policy.total_leaves if company_policy else 0
    sick_leave_balance = company_policy.sick_leaves if company_policy else 0
    paid_leave_balance = company_policy.paid_leaves if company_policy else 0
    pending_leaves = total_leaves_allowed - total_leaves_taken
    return render(request, 'leaves.html', {
        'employee': employee,
        'total_leaves_allowed': total_leaves_allowed,
        'sick_leave_balance': sick_leave_balance,
        'paid_leave_balance': paid_leave_balance,
        'leaves': leaves,
        'total_leaves_taken': total_leaves_taken,
        'pending_leaves': pending_leaves
    })


def approve_request(request, request_type='Pending'):
    # If the request method is not POST, render the approve_request.html template as before
    leave_requests = LeaveRequest.objects.filter(
        status=request_type).order_by('-request_time')
    paginator_leave = Paginator(leave_requests, 10)
    page_leave = request.GET.get('page', 1)
    print(page_leave)
    leave_requests = paginator_leave.page(page_leave)
    return render(request, 'approve_request.html', {'leave_requests': leave_requests, 'request_type': request_type})


def approve_request22(request, request_type='Pending'):
    attendance_requests = AttendanceRequest.objects.filter(
        status=request_type).order_by('-request_time')
    paginator_attendance = Paginator(attendance_requests, 10)
    page_attendance = request.GET.get('page', 1)
    attendance_requests = paginator_attendance.page(page_attendance)

    return render(request, 'approve_request22.html', {'attendance_requests': attendance_requests, 'request_type': request_type})


def company_holidays(request):
    holidays = Holiday.objects.all()
    return render(request, 'company_holidays.html', {"holidays": holidays})


def apply_attendance_rules(request):
    # Logging.objects.create(log="Running Cronjob")
    create_current_day_attendances()
    return HttpResponse('Creating todays attendance and marking absent on day before today')



def support(request):
    return render(request, 'support.html')


def contact(request):
    return render(request, 'contact.html')


def all_employees(request):
    # Default to current month if no parameters are provided
    current_year = datetime.datetime.now().year
    current_month = datetime.datetime.now().month

    return render(request, 'all_employees.html', {
        'current_year': current_year,
        'current_month': current_month
    })


def fetch_employee_data(request):
    year = int(request.GET.get('year', datetime.datetime.now().year))
    month = int(request.GET.get('month', datetime.datetime.now().month))

    # Calculate the first and last day of the month
    first_day = datetime.date(year, month + 1, 1)
    if month == 12:
        last_day = datetime.date(year + 1, 1, 1)
    else:
        last_day = datetime.date(year, month + 2, 1)

    print(first_day, last_day)

    employees = Attendance.objects.filter(
        date__gte=first_day, date__lt=last_day
    ).values(
        "employee__user__email",  # Grouping by employee email
        "employee__id",            # Also retrieving employee ID
        "employee__user__first_name",  # Also retrieving employee name
        "employee__user__last_name",
    ).annotate(
        year=ExtractYear('date'),  # Extract year from date
        month=ExtractMonth('date'),  # Extract month from date
        total_records=Count('id'),  # Count of total records for each employee
        present=Count(Case(When(is_present=True, then=Value(1)),
                      output_field=IntegerField())),  # Count of present days
        absent=Count(Case(When(is_absent=True, then=Value(1)),
                     output_field=IntegerField())),    # Count of absent days
        half_day=Count(Case(When(is_half_day=True, then=Value(
            1)), output_field=IntegerField())),  # Count of half-day attendance
        leave=Count(Case(When(is_leave=True, then=Value(1)),
                    output_field=IntegerField())),      # Count of leave days
        # Count of late attendance
        late=Count(Case(When(is_late=True, then=Value(1)),
                   output_field=IntegerField())),
        # Count of early departures
        early=Count(Case(When(is_early=True, then=Value(1)),
                    output_field=IntegerField())),
        # Sorting results by employee email
    ).order_by('employee__user__first_name')

    employee_data = list(employees)
    return JsonResponse({'employees': employee_data})


def approve_leave_request(request):
    if request.method == 'POST':
        print(request.POST)
        leave_request_id = request.POST.get('leave_request_id')
        approval_reason = request.POST.get('approveReason')
        try:
            leave_request = LeaveRequest.objects.get(id=leave_request_id)
        except LeaveRequest.DoesNotExist:
            messages.error(request, 'Leave request does not exist.')

        leave_request.status = "Approved"
        leave_request.reason = approval_reason
        leave_request.save()
        messages.success(
            request, 'Leave request has been approved successfully.')

        request_type = 'Pending'
        return redirect(reverse('approve_request', kwargs={'request_type': request_type}))


def reject_leave_request(request):
    if request.method == 'POST':
        leave_request_id = request.POST.get('leave_request_id')
        rejection_reason = request.POST.get('rejectReason')
        try:
            leave_request = LeaveRequest.objects.get(id=leave_request_id)
        except LeaveRequest.DoesNotExist:
            messages.error(request, 'Leave request does not exist.')

        leave_request.status = "Rejected"
        leave_request.reason = rejection_reason
        leave_request.save()
        messages.success(
            request, 'Leave request has been rejected successfully.')
        request_type = 'Pending'
        return redirect(reverse('approve_request', kwargs={'request_type': request_type}))


def approve_attendance_request(request):
    if request.method == 'POST':
        print(request.POST)
        attendance_request_id = request.POST.get('attendance_request_id')
        approval_reason = request.POST.get('approveReason')
        try:
            attendance_request = AttendanceRequest.objects.get(
                id=attendance_request_id)
        except AttendanceRequest.DoesNotExist:
            messages.error(request, 'Attendance request does not exist.')

        attendance_request.status = "Approved"
        attendance_request.reason = approval_reason
        attendance_request.save()

        att = Attendance.objects.get(
            employee=attendance_request.employee, date=attendance_request.date)
        if attendance_request.request_type == 'Check-in':
            att.clock_in = attendance_request.time
        else:
            att.clock_out = attendance_request.time
        att.save()

        messages.success(
            request, 'Attendance request has been approved successfully.')
        request_type = 'Pending'
        return redirect(reverse('approve_request22', kwargs={'request_type': request_type}))


def reject_attendance_request(request):
    if request.method == 'POST':
        attendance_request_id = request.POST.get('attendance_request_id')
        rejection_reason = request.POST.get('rejectReason')
        try:
            attendance_request = AttendanceRequest.objects.get(
                id=attendance_request_id)
        except AttendanceRequest.DoesNotExist:
            messages.error(request, 'Attendance request does not exist.')

        attendance_request.status = "Rejected"
        attendance_request.reason = rejection_reason
        attendance_request.save()
        messages.success(
            request, 'Attendance request has been rejected successfully.')
        request_type = 'Pending'
        return redirect(reverse('approve_request22', kwargs={'request_type': request_type}))


def generate_excel(request, emp_id):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    employee = Employee.objects.get(id=emp_id)
    response['Content-Disposition'] = f'attachment; filename="Report - {employee.user.first_name} {employee.user.last_name}.xlsx"'

    # Fetch data from the database and group by year and month
    attendances = Attendance.objects.filter(employee__id=emp_id).annotate(
        year=ExtractYear('date'),
        month=ExtractMonth('date')
    ).order_by('year', 'month', 'date')

    # Create a workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Define styles
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center')
    cell_alignment = Alignment(horizontal='center')

    # Group data by year and month
    grouped_data = {}
    for attendance in attendances:
        year = attendance.year
        month = attendance.month
        if (year, month) not in grouped_data:
            grouped_data[(year, month)] = []
        grouped_data[(year, month)].append({
            "date": attendance.date,
            "clock_in": attendance.clock_in if attendance.clock_in else "-",
            "clock_out": attendance.clock_out if attendance.clock_out else "-",
            "status": "Present" if attendance.is_present else "Absent" if attendance.is_absent else "Half" if attendance.is_half_day else "Leave"
        })

    # Create table data for each month
    row = 1
    for (year, month), records in grouped_data.items():
        month_name = calendar.month_name[month]
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=4)
        ws.cell(row=row, column=1).value = f'{month_name} {year}'
        ws.cell(row=row, column=1).font = header_font
        row += 1

        headers = ['Date', 'Clock In', 'Clock Out', 'Status']
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.alignment = header_alignment
        row += 1

        for record in records:
            ws.cell(row=row, column=1).value = str(record['date'])
            ws.cell(row=row, column=2).value = str(record['clock_in'])
            ws.cell(row=row, column=3).value = str(record['clock_out'])
            ws.cell(row=row, column=4).value = str(record['status'])

            for col_num in range(1, 5):
                ws.cell(row=row, column=col_num).alignment = cell_alignment

            row += 1

        row += 1  # Add some space between tables

    # Save the workbook to the response
    wb.save(response)
    return response


# Leave Setting
@login_required
def leave_setting(request):
    if request.method == 'POST':
        # Get the employee instance of the logged-in user
        employee = Employee.objects.get(user=request.user)
        company = employee.company

        early_late_mark_deduction = request.POST.get(
            'early_late_mark_deduction')
        late_deduction_after_days = request.POST.get(
            'late_deduction_after_days')
        early_deduction_after_days = request.POST.get(
            'early_deduction_after_days')
        consecutive_late_deduction_amount = request.POST.get(
            'consecutive_late_deduction_amount')
        monthly_late_deduction_amount = request.POST.get(
            'monthly_late_deduction_amount')
        weekly_late_deduction_amount = request.POST.get(
            'weekly_late_deduction_amount')
        monthly_leave = request.POST.get('monthly_leave')
        print(early_late_mark_deduction, late_deduction_after_days, early_deduction_after_days,
              consecutive_late_deduction_amount, monthly_late_deduction_amount, weekly_late_deduction_amount, monthly_leave)

        if not early_late_mark_deduction or not late_deduction_after_days or not early_deduction_after_days:
            messages.error(request, 'Please enter all fields.')
            return redirect('leave_setting')

        rules = CompanyRules.objects.get(company=company)
        rules.early_late_mark_deduction = early_late_mark_deduction
        rules.late_deduction_after_days = late_deduction_after_days
        rules.early_deduction_after_days = early_deduction_after_days
        rules.consecutive_late_deduction_amount = consecutive_late_deduction_amount
        rules.monthly_late_deduction_amount = monthly_late_deduction_amount
        rules.weekly_late_deduction_amount = weekly_late_deduction_amount
        rules.monthly_leave = monthly_leave

        rules.save()

        messages.success(request, 'Leave setting updated successfully.')

        return redirect('leave_setting')

    # Get the employee instance of the logged-in user
    employee = Employee.objects.get(user=request.user)
    company = employee.company

    rules = CompanyRules.objects.get(company=company)

    context = {
        'rules': rules
    }
    return render(request, 'leave_setting.html', context)


def penalities_setting(request):
    if request.method == 'POST':
        rules = request.POST.get('rules')
        period = request.POST.get('period')
        penalty = request.POST.get('penalty')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        created_at = timezone.now()
        updated_at = timezone.now()
        
        # Check if all fields are present
        if not rules or not period or not penalty or not start_date or not end_date:
            messages.error(request, 'Please enter all fields.')
            return redirect('penalities_setting')

        company = Employee.objects.get(user=request.user).company

        # Get or create the CompanyPenality instance
        instance, created = CompanyPenality.objects.get_or_create(
            company=company,
            start_date=start_date,
            end_date=end_date,
            defaults={
                'rules': rules,
                'period': period,
                'penalty': penalty,
                'created_at': created_at,
                'updated_at': updated_at
            }
        )
        if not created:
            instance.rules = rules
            instance.period = period
            instance.penalty = penalty
            instance.start_date = start_date
            instance.end_date = end_date
            instance.updated_at = updated_at
            instance.save()

        messages.success(request, 'Penalties updated successfully.')
        return redirect('penalities_setting')

    # Get the employee instance of the logged-in user
    employee = Employee.objects.get(user=request.user)
    company = employee.company
      
    # Get or initialize the CompanyPenality instance
    instance = CompanyPenality.objects.filter(company=company).order_by('-created_at').first()

    

    # Get companies and rule choices regardless of POST or initial GET request
    companies = Company.objects.all()
    rule_choices = CompanyPenality.RULE_CHOICES
    days_range = range(1, 31)  # Generate a list of days from 1 to 30

    return render(request, 'penalities_setting.html', {
        'companies': companies,
        'rule_choices': rule_choices,
        'instance': instance,
        'days_range': days_range,  # Pass the range to the template
    })

def forgot_password(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        token = get_random_string(length=32)
        PasswordResetRequest.objects.create(email=email, token=token)

        # Send password reset email
        subject = 'Password Reset'
        message = render_to_string(
            'email/password_reset_email.html', {'token': token})
        send_mail(subject, message, 'thenestiny@gmail.com', [email])

        return render(request, 'password_reset_request_sent.html')
    return render(request, 'forget_password.html')
